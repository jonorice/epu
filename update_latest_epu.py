#!/usr/bin/env python
"""Update the Irish main and domestic EPU workbooks.

This script automates the fallback workflow that is currently the most reliable:

1. Scrape Irish Times article counts for:
   - an overlap anchor month that already exists in the published output
   - the latest completed month
   - the current partial month through a supplied end date
2. Recover the combined-series raw share for the anchor month from the existing
   published output.
3. Extend the main series by applying Irish Times flagged-share growth from the
   anchor month to the completed and partial months.
4. Rebuild the domestic input dataset from the latest all-country file on
   policyuncertainty.com plus the updated Irish main series.
5. Recompute the domestic raw residual, then scale it back onto the published
   domestic-series scale using the overlapping historical output.

Example:
    python update_latest_epu.py --anchor-month 2026-01 --full-month 2026-02 --partial-end 2026-03-11
"""

from __future__ import annotations

import argparse
import calendar
import importlib.util
import shutil
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Dict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.decomposition import PCA
from sklearn.linear_model import LinearRegression


ROOT = Path(__file__).resolve().parent
SCRAPER_PATH = ROOT / "1m_individual_papers.py"
MAIN_WORKBOOK_PATH = ROOT / "main_epu_construction.xlsx"
DOMESTIC_INPUT_PATH = ROOT / "dataset_est_domestic.xlsx"
PUBLIC_OUTPUT_PATH = ROOT / "Ireland_Policy_Uncertainty_Data_Rice.xlsx"
PAPER_URL = "https://www.esr.ie/article/view/2531"
STATE_DB_NAME = "irish_epu_state.sqlite"


@dataclass(frozen=True)
class MonthCounts:
    month: pd.Timestamp
    flagged: int
    total: int

    @property
    def share(self) -> float:
        if self.total <= 0:
            raise ValueError(f"Month {self.month:%Y-%m} has a non-positive total.")
        return self.flagged / self.total


def parse_month(value: str) -> pd.Timestamp:
    try:
        return pd.Timestamp(f"{value}-01")
    except Exception as exc:  # pragma: no cover - defensive parsing
        raise argparse.ArgumentTypeError(
            f"Invalid month '{value}'. Use YYYY-MM."
        ) from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--anchor-month",
        required=True,
        type=parse_month,
        help="Existing full month in the public output used as the combined-series anchor, format YYYY-MM.",
    )
    parser.add_argument(
        "--full-month",
        required=True,
        type=parse_month,
        help="Latest completed month to estimate from Irish Times growth, format YYYY-MM.",
    )
    parser.add_argument(
        "--partial-end",
        required=True,
        type=pd.Timestamp,
        help="End date for the current partial month, format YYYY-MM-DD.",
    )
    parser.add_argument(
        "--scratch-dir",
        default=str(ROOT / "_scratch_epu"),
        help="Directory used for resumable Irish Times scrape state.",
    )
    parser.add_argument(
        "--keep-scratch",
        action="store_true",
        help="Keep the scrape state directory after a successful run.",
    )
    return parser.parse_args()


def month_end(month: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(
        year=month.year,
        month=month.month,
        day=calendar.monthrange(month.year, month.month)[1],
    )


def load_scraper_module():
    spec = importlib.util.spec_from_file_location("irish_epu_scraper", SCRAPER_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load scraper module from {SCRAPER_PATH}.")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def sql_frame(conn: sqlite3.Connection, query: str, params: tuple = ()) -> pd.DataFrame:
    return pd.read_sql_query(query, conn, params=params)


def run_irish_times_counts(
    scraper,
    month: pd.Timestamp,
    end_date: pd.Timestamp,
    scratch_root: Path,
) -> MonthCounts:
    if month.to_period("M") != end_date.to_period("M"):
        raise ValueError(
            f"Month {month:%Y-%m} and end date {end_date:%Y-%m-%d} are not in the same month."
        )

    outdir = scratch_root / f"irish_times_{month:%Y_%m}"
    cache_dir = outdir / "cache"
    state_dir = outdir / "state"
    for path in (outdir, cache_dir / "sitemaps", cache_dir / "articles", cache_dir / "robots", state_dir):
        path.mkdir(parents=True, exist_ok=True)

    start = month.date()
    end = end_date.date()

    conn = scraper.db_connect(state_dir / STATE_DB_NAME)
    try:
        scraper.db_init(conn)
        session = scraper.build_session(None)
        scraper.discover_irish_times(session, conn, cache_dir, start, end)

        while True:
            pending = sql_frame(
                conn,
                """
                SELECT COUNT(*) AS pending_n
                FROM articles
                WHERE source = 'IrishTimes'
                  AND month = ?
                  AND (
                    status = 'discovered'
                    OR (status = 'fetch_failed' AND attempts_fetch_fail < ?)
                  );
                """,
                (month.date().isoformat(), int(scraper.MAX_ATTEMPTS_FETCH_FAIL)),
            )
            pending_n = int(pending["pending_n"].iloc[0])
            if pending_n == 0:
                break

            processed = scraper.fetch_and_process_pending(
                conn,
                {"IrishTimes": session},
                cache_dir,
                start,
                end,
                ["IrishTimes"],
            )
            if processed == 0:
                raise RuntimeError(
                    f"Irish Times scrape stalled for {month:%Y-%m} with {pending_n} pending rows."
                )

        counts = sql_frame(
            conn,
            """
            SELECT
              COUNT(*) AS total_articles,
              SUM(COALESCE(match, 0)) AS matched_articles
            FROM articles
            WHERE source = 'IrishTimes'
              AND month = ?
              AND status != 'out_of_range';
            """,
            (month.date().isoformat(),),
        )
        total = int(counts["total_articles"].iloc[0])
        flagged = int(counts["matched_articles"].iloc[0] or 0)
        return MonthCounts(month=month, flagged=flagged, total=total)
    finally:
        conn.close()


def load_public_output() -> pd.DataFrame:
    df = pd.read_excel(PUBLIC_OUTPUT_PATH)
    df = df[pd.to_datetime(df["Date"], errors="coerce").notna()].copy()
    df["Date"] = pd.to_datetime(df["Date"])
    return df


def load_main_workbook() -> pd.DataFrame:
    raw = pd.read_excel(MAIN_WORKBOOK_PATH)
    return raw.rename(
        columns={
            raw.columns[0]: "Date",
            raw.columns[1]: "Flagged",
            raw.columns[2]: "Total",
            raw.columns[3]: "Share",
            raw.columns[4]: "Ireland_main_epu",
        }
    )


def recover_anchor_share(
    main_df: pd.DataFrame,
    public_output_df: pd.DataFrame,
    anchor_month: pd.Timestamp,
) -> float:
    main_df = main_df.copy()
    main_df["Date"] = pd.to_datetime(main_df["Date"])

    overlap = main_df[["Date", "Share"]].merge(
        public_output_df[["Date", "Ireland_main_epu"]],
        on="Date",
        how="inner",
    )
    overlap = overlap[overlap["Share"].notna()].copy()
    if overlap.empty:
        raise RuntimeError("No overlap between the main workbook and public output.")

    ratio = (overlap["Ireland_main_epu"] / overlap["Share"]).mean()
    anchor_row = public_output_df.loc[public_output_df["Date"] == anchor_month]
    if anchor_row.empty:
        raise RuntimeError(f"Anchor month {anchor_month:%Y-%m} is not present in {PUBLIC_OUTPUT_PATH.name}.")
    return float(anchor_row["Ireland_main_epu"].iloc[0] / ratio)


def update_main_series(
    main_df: pd.DataFrame,
    replacement_shares: Dict[pd.Timestamp, float],
) -> pd.DataFrame:
    updated = main_df.copy()
    updated["Date"] = pd.to_datetime(updated["Date"])
    updated = updated[["Date", "Flagged", "Total", "Share"]]

    for month, share in replacement_shares.items():
        mask = updated["Date"] == month
        if mask.any():
            updated.loc[mask, ["Flagged", "Total", "Share"]] = [np.nan, np.nan, share]
        else:
            updated = pd.concat(
                [
                    updated,
                    pd.DataFrame(
                        [{"Date": month, "Flagged": np.nan, "Total": np.nan, "Share": share}]
                    ),
                ],
                ignore_index=True,
            )

    updated = updated.sort_values("Date").reset_index(drop=True)
    mean_share = updated["Share"].mean()
    updated["Ireland_main_epu"] = updated["Share"] / mean_share * 100
    return updated


def build_domestic_input(updated_main: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    local_dom = pd.read_excel(DOMESTIC_INPUT_PATH)
    local_dom["month"] = pd.to_datetime(local_dom["month"])

    web = pd.read_excel("https://www.policyuncertainty.com/media/All_Country_Data.xlsx")
    web = web[pd.to_numeric(web["Year"], errors="coerce").notna()].copy()
    web["Date"] = pd.to_datetime(
        dict(year=web["Year"].astype(int), month=web["Month"].astype(int), day=1)
    )
    web = web[web["Date"] >= "1997-01-01"].copy()
    web["Hybrid China"] = ((web["SCMP China"] + web["Mainland China"]) / 2).where(
        web["SCMP China"].notna() & web["Mainland China"].notna(),
        web["China"],
    )

    if "Pakistan" not in local_dom.columns:
        pak_hist = web[["Date", "Pakistan"]].rename(columns={"Date": "month"})
        local_dom = local_dom.drop(columns=["Netherlands"], errors="ignore").merge(
            pak_hist,
            on="month",
            how="left",
        )
    else:
        local_dom = local_dom.drop(columns=["Netherlands"], errors="ignore")

    append = web[web["Date"] > local_dom["month"].max()].copy()
    append_dom = pd.DataFrame(
        {
            "month": append["Date"],
            "Australia": append["Australia"],
            "Brazil": append["Brazil"],
            "Canada": append["Canada"],
            "Chile": append["Chile"],
            "Hybrid China": append["Hybrid China"],
            "France": append["France"],
            "Germany": append["Germany"],
            "Greece": append["Greece"],
            "India": append["India"],
            "Italy": append["Italy"],
            "Japan": append["Japan"],
            "Korea": append["Korea"],
            "Russia": append["Russia"],
            "Spain": append["Spain"],
            "UK": append["UK"],
            "US": append["US"],
            "Pakistan": append["Pakistan"],
            "Ireland_Rice": np.nan,
        }
    )

    dom_input = pd.concat([local_dom, append_dom], ignore_index=True)
    main_for_dom = updated_main[["Date", "Ireland_main_epu"]].copy()
    dom_input = dom_input.rename(columns={"month": "Date"}).merge(
        main_for_dom,
        on="Date",
        how="left",
    )
    dom_input["Ireland_Rice"] = dom_input["Ireland_main_epu"]
    dom_input = dom_input.drop(columns=["Ireland_main_epu"])
    dom_input = dom_input[dom_input["Date"] <= web["Date"].max()].copy()
    return dom_input, web


def compute_domestic_series(
    dom_input: pd.DataFrame,
    web: pd.DataFrame,
    public_output_df: pd.DataFrame,
) -> pd.DataFrame:
    foreign_cols = [
        "Australia",
        "Brazil",
        "Canada",
        "Chile",
        "Hybrid China",
        "France",
        "Germany",
        "Greece",
        "India",
        "Italy",
        "Japan",
        "Korea",
        "Russia",
        "Spain",
        "UK",
        "US",
        "Pakistan",
    ]

    pca_base = dom_input[["Date"] + foreign_cols].merge(
        web[["Date", "GEPU_current"]],
        on="Date",
        how="left",
    )
    X = pca_base[foreign_cols].to_numpy(dtype=float)
    X_std = (X - X.mean(axis=0)) / X.std(axis=0, ddof=1)
    pca = PCA()
    scores = pca.fit_transform(X_std)
    prop = pca.explained_variance_ratio_
    num_components = int(np.argmax(np.cumsum(prop) >= 0.7) + 1)

    weighted = scores[:, :num_components] @ prop[:num_components]
    weighted_std = (weighted - weighted.mean()) / weighted.std(ddof=1)
    pca_base["Weighted_PCA"] = (
        weighted_std * pca_base["GEPU_current"].std(ddof=1) + pca_base["GEPU_current"].mean()
    )
    for lag in [1, 2, 3]:
        pca_base[f"Lag{lag}"] = pca_base["Weighted_PCA"].shift(lag)

    reg = pca_base.merge(
        dom_input[["Date", "Ireland_Rice"]],
        on="Date",
        how="left",
    ).dropna(subset=["Ireland_Rice", "Lag1", "Lag2", "Lag3"])
    raw_model = LinearRegression().fit(
        reg[["Weighted_PCA", "Lag1", "Lag2", "Lag3"]],
        reg["Ireland_Rice"],
    )
    reg["raw_domestic"] = reg["Ireland_Rice"] - raw_model.predict(
        reg[["Weighted_PCA", "Lag1", "Lag2", "Lag3"]]
    )

    published_dom = public_output_df[["Date", "Ireland_domestic_uncertainty"]].copy()
    published_dom = published_dom[published_dom["Ireland_domestic_uncertainty"].notna()]
    scale_fit = reg.merge(published_dom, on="Date", how="inner")
    scale_model = LinearRegression().fit(
        scale_fit[["raw_domestic"]],
        scale_fit["Ireland_domestic_uncertainty"],
    )
    reg["Ireland_domestic_uncertainty"] = scale_model.predict(reg[["raw_domestic"]])
    return reg[["Date", "Ireland_domestic_uncertainty"]].copy()


def write_main_workbook(
    updated_main: pd.DataFrame,
    anchor_month: pd.Timestamp,
    full_month: pd.Timestamp,
    partial_end: pd.Timestamp,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        None,
        "Count of flagged articles (Irish Times, Irish Independent, Irish Examiner)",
        "Total Articles (Irish Times, Irish Independent, Irish Examiner)",
        "Share of flagged articles (Irish Times, Irish Independent, Irish Examiner)",
        "Rescaled FINAL EPU",
    ]
    for column, value in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=value)

    for row_n, row in enumerate(updated_main.itertuples(index=False), start=2):
        ws.cell(row=row_n, column=1, value=row.Date.to_pydatetime())
        if pd.notna(row.Flagged):
            ws.cell(row=row_n, column=2, value=float(row.Flagged))
        if pd.notna(row.Total):
            ws.cell(row=row_n, column=3, value=float(row.Total))
        ws.cell(row=row_n, column=4, value=float(row.Share))
        ws.cell(row=row_n, column=5, value=float(row.Ireland_main_epu))

    last_row = len(updated_main) + 1
    ws["F1"] = "Note"
    ws["F2"] = f"Anchor month for the Irish Times growth fallback: {anchor_month:%Y-%m}."
    ws["F3"] = f"Latest completed month updated from fallback growth: {full_month:%Y-%m}."
    ws["F4"] = f"Partial month updated through {partial_end:%Y-%m-%d}."
    ws["F5"] = "Rows overwritten via fallback leave combined flagged/total counts blank when direct combined counts are unavailable."
    ws["G11"] = "RESCALING"
    ws["H12"] = "old"
    ws["I12"] = "new"
    ws["G13"] = "mean"
    ws["H13"] = f"=AVERAGE(D2:D{last_row})"
    ws["I13"] = 100
    ws["G14"] = "stdev"
    ws["H14"] = f"=STDEV.P(D2:D{last_row})"
    ws["I14"] = "=(H14/H13)*I13"
    wb.save(MAIN_WORKBOOK_PATH)


def write_domestic_input(dom_input: pd.DataFrame) -> None:
    to_write = dom_input.rename(columns={"Date": "month"})[
        [
            "month",
            "Australia",
            "Brazil",
            "Canada",
            "Chile",
            "Hybrid China",
            "France",
            "Germany",
            "Greece",
            "India",
            "Italy",
            "Japan",
            "Korea",
            "Russia",
            "Spain",
            "UK",
            "US",
            "Pakistan",
            "Ireland_Rice",
        ]
    ].copy()
    with pd.ExcelWriter(DOMESTIC_INPUT_PATH, engine="openpyxl") as writer:
        to_write.to_excel(writer, index=False, sheet_name="Sheet1")


def write_public_output(
    updated_main: pd.DataFrame,
    domestic_df: pd.DataFrame,
) -> None:
    output = updated_main[["Date", "Ireland_main_epu"]].merge(
        domestic_df,
        on="Date",
        how="left",
    )
    output = output.sort_values("Date").reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Ireland_main_epu")
    ws.cell(row=1, column=3, value="Ireland_domestic_uncertainty")

    for row_n, row in enumerate(output.itertuples(index=False), start=2):
        ws.cell(row=row_n, column=1, value=row.Date.to_pydatetime())
        ws.cell(row=row_n, column=2, value=float(row.Ireland_main_epu))
        if pd.notna(row.Ireland_domestic_uncertainty):
            ws.cell(row=row_n, column=3, value=float(row.Ireland_domestic_uncertainty))

    cite_row = len(output) + 3
    ws.cell(
        row=cite_row,
        column=1,
        value=(
            'Please cite: Rice, J. "Economic Policy Uncertainty in Small Open Economies: '
            'A Case Study of Ireland." The Economic and Social Review, Vol. 54 No. 4, '
            "Winter 2023. These data can be used freely with attribution to the author and paper."
        ),
    )
    link_row = cite_row + 1
    ws.cell(row=link_row, column=1, value="Paper link")
    link_cell = ws.cell(row=link_row, column=2, value=PAPER_URL)
    link_cell.hyperlink = PAPER_URL
    link_cell.style = "Hyperlink"
    link_cell.font = Font(color="0563C1", underline="single")
    wb.save(PUBLIC_OUTPUT_PATH)


def main() -> None:
    args = parse_args()
    scratch_root = Path(args.scratch_dir)
    scratch_root.mkdir(parents=True, exist_ok=True)

    partial_month = pd.Timestamp(args.partial_end).to_period("M").to_timestamp()
    if args.anchor_month >= args.full_month:
        raise SystemExit("The anchor month must be earlier than the full month.")
    if args.full_month >= partial_month:
        raise SystemExit("The full month must be earlier than the partial month.")

    scraper = load_scraper_module()
    anchor_counts = run_irish_times_counts(scraper, args.anchor_month, month_end(args.anchor_month), scratch_root)
    full_counts = run_irish_times_counts(scraper, args.full_month, month_end(args.full_month), scratch_root)
    partial_counts = run_irish_times_counts(scraper, partial_month, pd.Timestamp(args.partial_end), scratch_root)

    public_output_df = load_public_output()
    main_df = load_main_workbook()
    main_df["Date"] = pd.to_datetime(main_df["Date"])

    anchor_share = recover_anchor_share(main_df, public_output_df, args.anchor_month)
    full_share = anchor_share * (full_counts.share / anchor_counts.share)
    partial_share = anchor_share * (partial_counts.share / anchor_counts.share)

    updated_main = update_main_series(
        main_df,
        {
            args.full_month: full_share,
            partial_month: partial_share,
        },
    )
    dom_input, web = build_domestic_input(updated_main)
    domestic_df = compute_domestic_series(dom_input, web, public_output_df)

    write_main_workbook(updated_main, args.anchor_month, args.full_month, pd.Timestamp(args.partial_end))
    write_domestic_input(dom_input)
    write_public_output(updated_main, domestic_df)

    print(f"Anchor month Irish Times counts: {anchor_counts.flagged} flagged / {anchor_counts.total} total")
    print(f"Full month Irish Times counts: {full_counts.flagged} flagged / {full_counts.total} total")
    print(
        f"Partial month Irish Times counts through {pd.Timestamp(args.partial_end):%Y-%m-%d}: "
        f"{partial_counts.flagged} flagged / {partial_counts.total} total"
    )
    print(
        updated_main.loc[
            updated_main["Date"].isin([args.full_month, partial_month]),
            ["Date", "Ireland_main_epu"],
        ].to_string(index=False)
    )

    if not args.keep_scratch:
        shutil.rmtree(scratch_root, ignore_errors=True)


if __name__ == "__main__":
    main()
